import pyodbc
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font

conn = pyodbc.connect(
    'Driver={IBM i Access ODBC Driver}; System=localhost; '
    'UID=osswxx; PWD=osswxx;'
    )
cursor = conn.cursor()
df = pd.read_sql_query("select * from qeol.tokmsp", conn)
df2 = pd.read_sql_query("""select length,numeric_scale,numeric_precision,
    column_heading from qsys2.syscolumns
    where table_name = 'TOKMSP' and table_schema = 'QEOL'""", conn)
cursor.close()
conn.close()

df = df.applymap(lambda x: x.rstrip(" ""　") if isinstance(x, str) else x)
header = tuple(df2["COLUMN_HEADING"]. \
    apply(lambda x: x.replace(" ","").replace("　","")))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "得意先マスター"

for i, value in enumerate(header, start=1):
    ws.cell(1, i).value = value

rows = dataframe_to_rows(df, index=False, header=True)
for row_no, row in enumerate(rows, start=2):
    for col_no, value in enumerate(row, start=1):
        ws.cell(row_no, col_no).value = value

for i, column_cells in enumerate(ws.columns):
    if df2["NUMERIC_SCALE"][i] == None:
        width = df2["LENGTH"][i]
        if width > 60: width = 60 
    else:
         width = df2["LENGTH"][i] + 4
    if width < 12: width = 12 
    ws.column_dimensions[get_column_letter(i + 1)].width = width

table = Table(displayName='テーブル1', ref="A2:" + \
    get_column_letter(ws.max_column) + str(ws.max_row))
table.tableStyleInfo = TableStyleInfo(name='TableStyleMedium9', showRowStripes=True)
ws.add_table(table)

font = Font(name='メイリオ')
for row in ws:
    for cell in row:
        ws[cell.coordinate].font = font
ws.sheet_view.zoomScale = 75
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.fitToHeight = False
ws.sheet_view.view = "pageBreakPreview" 

wb.save('./tokmsp2.xlsx')
wb.close()